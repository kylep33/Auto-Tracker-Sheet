import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import re


def load_excel_file(file_path):
    try:
        # Load the Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_path)
        return workbook
    except InvalidFileException:
        print(f"Error: Invalid file '{file_path}'")
        exit(-1)


def read_title(sheet):
    # Extract address from cell A2
    title = sheet['A2'].value
    return title


def read_job(sheet):
    # Assuming title is in the third row (A3)
    job = sheet['A1'].value
    return job


def parse_title_text(input_text):
    # Define a regular expression pattern to extract the controller type and unit type
    pattern = r'([^\]]*?)\s*Controller Points List for\s*([^\(]*?)\(Typ\. of (\d+)\)'

    # Use re.search to find the match
    match = re.search(pattern, input_text)

    if match:
        # Extract matched groups
        controller_type = match.group(1).strip()
        unit_type = match.group(2).strip()
        num_of_units = match.group(3)
    else:
        # Set defaults if no match is found
        controller_type = "controller type here"
        unit_type = "unit type here"
        num_of_units = "1"

    return {
        'controller_type': controller_type,
        'unit_type': unit_type,
        'num_of_units': num_of_units
    }

def extract_unit_type(title):
    # Dictionary used to be able to extract from title. the words on the right are what can be in the title
    reference_dict = {
        "EF": ["EF", "Exhaust", "Exhaust Fan"],
        "WSHP": ["WSHP", "WATER SOURCE HEAT PUMP"],
        "VAV": ["VAV"],
        "FC": ["Fan Coil", "FC"],
        "ASHP": ["ASHP", "Air Source Heat Pump"],
        "VRF": ["VRF"],
        "Package Unit": ["Package Unit", "PU"],
        "Mini Split": ["Mini Split", "MS"],
        "Ducted Split": ["Ducted"]
    }

    found_terms = []

    # Loop through Reference Dictionary, as seen above
    for key, targets in reference_dict.items():
        for target in targets:
            if target.lower() in title.lower():
                found_terms.append(key)
                break  # Break the inner loop if a match is found for the current key

    if found_terms:
        return found_terms

    # If no term from reference_dict is found, try to extract between "List for " and "(Typ."
    start_marker = "for "
    end_marker = "(T"
    start_index = title.lower().find(start_marker.lower())
    end_index = title.lower().find(end_marker.lower())

    if start_index != -1 and end_index != -1:
        unit_type = title[start_index + len(start_marker):end_index].strip()
        return unit_type

    # If neither reference_dict terms nor "List for ... (Typ." is found, return "enter unit type"
    return "enter unit type here"


def display_sheet_contents(sheet):
    # Display the contents of the Excel sheet
    print("Contents of the Excel sheet:")
    for row in sheet.iter_rows(values_only=True):
        print(row)


def create_ip_op_dict(sheet):
    """This function locates all the rows that contain IP and OP and then uses those rows to build the IP OP dictionary"""
    relevant_rows = _read_rows_with_ip_op(sheet)
    ip_op_dict = _create_dictionary_from_rows(relevant_rows)
    return ip_op_dict


def _read_rows_with_ip_op(sheet):
    # Read rows that contain an element with "IP#" or "OP#"
    relevant_rows = []

    for row in sheet.iter_rows(values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and ("IP" in cell_value or "OP" in cell_value):
                relevant_rows.append(row)
                break  # Break the inner loop if a relevant element is found in the row

    return relevant_rows


def _create_dictionary_from_rows(relevant_rows):
    ip_op_dict = {}

    for row in relevant_rows:
        for i, cell_value in enumerate(row):
            if cell_value is not None and ('IP' in cell_value or 'OP' in cell_value):
                ip_key = cell_value
                ip_values = row[i + 1:i + 3]  # Assuming the next two elements are relevant
                ip_op_dict[ip_key] = ip_values

    return ip_op_dict


def main():
    # Example usage
    file_path = r"C:\Users\delta\PycharmProjects\Project Tracking Excel Sheet\testing_dir\Points List Template.xlsx"
    sheet_name = "DAC-304 EF-6"  # Replace with the actual sheet name

    # Load the Excel file
    workbook = load_excel_file(file_path)

    # Access the desired sheet
    sheet = workbook[sheet_name]

    display_sheet_contents(sheet)
    print()

    # Read and display title
    # job = read_job(sheet)
    # print("\njob:")
    # print(job)
    #
    # Read and display address
    title = read_title(sheet)
    print("\ntitle:")
    print(title)

    result = parse_title_text(title)

    if result:
        print("Controller Type:", result['controller_type'])
        print("Unit Type:", result['unit_type'])
        print("Number of Units:", result['num_of_units'])
    else:
        print("No match found.")
    print(extract_unit_type(title))

    print(create_ip_op_dict(sheet))


if __name__ == "__main__":
    main()
