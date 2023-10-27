import os

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

FULL_EXCEL_PATH = ''


def create_excel_workbook(job_name, unit_type, target_directory):
    # Create the full path for the target file
    file_name = f"Project Tracker - {job_name}.xlsx"
    full_path = os.path.join(target_directory, file_name)

    return full_path

def create_unit_sheets(workbook, unit_type):
    sheet1 = workbook.active

    # Otherwise, create a new sheet with the specified name
    workbook.create_sheet(title=f"Install - {unit_type}")

    # Create the second sheet named "Engineering Startup"
    workbook.create_sheet(title=f"Engineering Startup - {unit_type}")

def close_workbook(workbook):
    workbook.close()


def build_workbook(workbook, full_path, job_name, unit_type, number_of_units, ip_op_dict):
    build_install_sheet(workbook, full_path, unit_type, number_of_units, ip_op_dict)
    build_startup_sheet(workbook, full_path, job_name, unit_type, number_of_units, ip_op_dict)


def build_install_sheet(workbook, full_path, unit_type, number_of_units, ip_op_dict):

    install_sheet = find_install_sheet(workbook, unit_type)
    if not install_sheet:
        install_sheet = workbook.create_sheet(title=f"Install - {unit_type}")

    # Creating correct columns
    _set_headers_in_install_sheet(install_sheet, full_path)
    _create_ip_op_columns(install_sheet, ip_op_dict, "install")
    _add_end_rows(install_sheet)

    # Adding units
    add_units_to_sheet(install_sheet, unit_type, number_of_units)

    # Formatting stuff
    resize_install_sheet(install_sheet)
    change_colors_install(install_sheet)
    freeze_row_and_column(install_sheet, 1, 'A')
    center_all_cells(install_sheet)

    workbook.save(full_path)


def find_install_sheet(workbook, unit_type):
    # Find the sheet with a name starting with "Install -"
    install_sheet = None
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(f"Install - {unit_type}"):
            print(sheet_name)
            install_sheet = workbook[sheet_name]
            return install_sheet


def find_startup_sheet(workbook, unit_type):
    # Find the sheet with a name starting with "Startup -"
    startup_sheet = None
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(f"Engineering Startup - {unit_type}"):
            print(sheet_name)
            startup_sheet = workbook[sheet_name]
            return startup_sheet


def center_all_cells(sheet):
    # Iterate through all cells in the sheet except e1-g1
    for row in sheet.iter_rows():
        for cell in row:
            if cell.coordinate != "D1" and cell.coordinate != "E1" and cell.coordinate != "F1" and cell.coordinate != "G1":
                cell.alignment = Alignment(horizontal='center', vertical='center')


def add_units_to_sheet(sheet, unit_type, number_of_units):
    # Check if there are enough rows for the units, if not, insert rows

    # Assuming unit_names is a list of unit names
    for i in range(number_of_units):
        row_index = i + 2  # Start at row 2
        sheet.cell(row=row_index, column=3, value=unit_type)


def _set_headers_in_install_sheet(install_sheet, full_path):
    # Find the sheet with a name starting with "Install -"

    headers = ["UNIT#", "ADD#", "UNIT TYPE", "Fully\ninstalled", "T-Stat\nLabeled", "T-Grid\nLabeled", "Controller\ninstalled"]
    if install_sheet:
        # Iterate over the headers and set them in the first row of the found sheet
        for col_num, header in enumerate(headers, 1):
            cell = install_sheet.cell(row=1, column=col_num)
            # Set the header value and alignment to wrap text
            cell.value = header
            cell.alignment = Alignment(wrapText=True)

        # Save the workbook with the updated headers


def _set_headers_in_startup_sheet(startup_sheet, job_name):
    # Leave line 2 and 3 blank

    # Leave line 4 blank for now
    # ...

    # Headers on line 5
    headers = ["UNIT#", "ADD#", "UNIT TYPE"]
    for col_num, header in enumerate(headers, 1):
        cell = startup_sheet.cell(row=5, column=col_num)
        cell.value = header
        cell.alignment = Alignment(wrapText=True)


def _create_ip_op_columns(install_sheet, ip_op_dict, sheet_type):
    # differences between install and startup
    if sheet_type == "install":
        desired_row = 1
        max_col = install_sheet.max_column
    else:
        desired_row = 5
        max_col = 3

    purple_fill = PatternFill(start_color="CC99CC", end_color="CC99CC", fill_type="solid")

    # Extract keys that are not "Spare" and split them into IP and OP categories
    ip_keys = [key for key in ip_op_dict if key.startswith('IP') and "Spare" not in ip_op_dict[key][1]]
    op_keys = [key for key in ip_op_dict if key.startswith('OP') and "Spare" not in ip_op_dict[key][1]]

    # Combine IP and OP keys, inserting IP keys first
    ordered_keys = ip_keys + op_keys

    # Insert the columns into the sheet starting from the farthest right empty column
    for col_num, key in enumerate(ordered_keys, 1):
        header = shorten_ip_op(ip_op_dict[key][1])
        col_to_insert = max_col + col_num
        install_sheet.insert_cols(col_to_insert)
        cell = install_sheet.cell(row=desired_row, column=col_to_insert, value=header)
        cell.fill = purple_fill
        cell.alignment = Alignment(wrapText=True)


def shorten_ip_op(header):
    header_dict = {
        'RAT': ['Return Air Temp', 'RA Temp'],
        'SAT': ['Supply Air Temp', 'SA Temp'],
        'Spce Tmp': ['Space Temp'],
        'Wall Tmp': ['Wall Module Temp'],
        'Tmp': ['Temperature'],
        'eZNS tmp': ['eZNS Space Temp'],
        'Dis Air Tmp': ['Discharge Air Temp']
    }

    # Convert the input header to lowercase for case-insensitive comparison
    header = header.lower()

    for key, values in header_dict.items():
        for value in values:
            if value.lower() in header:
                return key

    return header



def _add_end_rows(install_sheet):
    # Find the last used column number in the sheet
    max_col = install_sheet.max_column

    # Insert two columns at the very end of the sheet
    install_sheet.insert_cols(max_col + 1, amount=2)

    # Set the titles for the added columns
    install_sheet.cell(row=1, column=max_col + 1, value="Installer Initials")
    install_sheet.cell(row=1, column=max_col + 2, value="               NOTES               ")

    # Print information to indicate that the operation is complete
    print("Installer Initials and NOTES rows added to the very end.")


def freeze_row_and_column(sheet, row, col):
    # Define the row and column to freeze (1 for the first row, 'A' for the first column)

    # Create a pane that freezes the specified row and column
    sheet.freeze_panes = f"{col}{row + 1}"


def resize_install_sheet(sheet):
    change_all_column_width(sheet)
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10


def change_all_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width


def change_colors_install(sheet):
    # Define colors
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    purple_fill = PatternFill(start_color="CC99CC", end_color="CC99CC", fill_type="solid")
    blue_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    grey_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Apply specific formatting
    for column in ['A', 'B', 'C']:
        for row in range(1, 2):  # Assuming you are working with the first row
            sheet[column + str(row)].fill = red_fill

    sheet['D1'].fill = yellow_fill

    for column in ['E', 'F', 'G']:
        for row in range(1, 2):
            sheet[column + str(row)].fill = green_fill

    sheet['H1'].fill = purple_fill

    for cell in sheet[1]:
        if cell.value is not None and cell.value.startswith("IP"):
            cell.fill = purple_fill
        elif cell.value is not None and cell.value.startswith("OP"):
            cell.fill = blue_fill
        elif cell.value is not None and cell.value.startswith("Installer"):
            cell.fill = yellow_fill
        elif cell.value is not None and cell.value.startswith("               NOTES"):
            cell.fill = grey_fill


def change_colors_startup(sheet):
    # Define colors
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    greener_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")

    purple_fill = PatternFill(start_color="CC99CC", end_color="CC99CC", fill_type="solid")
    blue_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    grey_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Apply specific formatting
    for column in ['A', 'B', 'C']:
        for row in range(5, 6):  # Assuming you are working with the first row
            sheet[column + str(row)].fill = red_fill

    for cell in sheet[5]:
        if cell.value is not None and cell.value.startswith("IP"):
            cell.fill = purple_fill
        elif cell.value is not None and cell.value.startswith("OP"):
            cell.fill = blue_fill
        elif cell.value is not None and cell.value.startswith("Min CFM"):
            cell.fill = green_fill
        elif cell.value is not None and cell.value.startswith("Max CFM"):
            cell.fill = green_fill
        elif cell.value is not None and cell.value.startswith("Engineer"):
            cell.fill = greener_fill
        elif cell.value is not None and cell.value.startswith("               NOTES"):
            cell.fill = grey_fill


def _____________________STARTUP_SHEET______________________():
    return


def build_startup_sheet(workbook, full_path, job_name, unit_type, number_of_units, ip_op_dict):
    startup_sheet = find_startup_sheet(workbook, unit_type)

    # Creating correct columns
    _set_headers_in_startup_sheet(startup_sheet, job_name)
    _create_ip_op_columns(startup_sheet, ip_op_dict, "startup")

    add_heating_cooling_headers(startup_sheet, 5,unit_type)
    _add_end_rows_startup(startup_sheet, 5,unit_type)
    #
    # # Adding units
    add_units_to_sheet_startup(startup_sheet, unit_type, number_of_units)
    #
    # # Formatting stuff
    # insert_image_into_sheet('ac_logo_for_startup.jpg', startup_sheet)
    resize_startup_sheet(startup_sheet)
    _set_title_in_startup_sheet(startup_sheet, job_name)
    change_colors_startup(startup_sheet)
    freeze_row_and_column(startup_sheet, 5, 'A')
    center_all_cells(startup_sheet)

    workbook.save(full_path)


def add_heating_cooling_headers(startup_sheet, row,unit_type):

    if unit_type == "VAV":
        headers = ["Min STP", "Max STP", "Room STP", "SA Temp", "Room Temp", "Pass"]
    else:
        headers = ["Room Set Point", "SA Temp", "Room Temp", "Pass"]

    title_headers = ["Cooling Test", "Heating Test"]
    title_fills = [PatternFill(start_color="AACCEE", end_color="AACCEE", fill_type="solid"),
                   PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")]
    content_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    for i in range(2):
        empty_col = get_first_empty_col(startup_sheet, row)
        title = title_headers[i]
        title_fill = title_fills[i]

        for col_num, header in enumerate(headers, empty_col):
            cell = startup_sheet.cell(row=row, column=col_num)
            cell.value = header
            cell.alignment = Alignment(wrapText=True)
            cell.fill = content_fill

        # Merge cells for the title row
        start_col = empty_col
        end_col = start_col + len(headers) - 1
        title_cell = startup_sheet.cell(row=row - 1, column=start_col)
        title_cell.value = title
        title_cell.fill = title_fill
        startup_sheet.merge_cells(start_row=row - 1, start_column=start_col, end_row=row - 1, end_column=end_col)


def get_first_empty_col(sheet, row):
    value_counter = 1
    for cell in sheet[row]:
        if cell.value is not None:
            value_counter = value_counter + 1
    return value_counter


def _add_end_rows_startup(startup_sheet, row, unit_type):
    # Find the last used column number in the sheet
    max_col = get_first_empty_col(startup_sheet, row) - 1

    # Insert two columns at the very end of the sheet
    startup_sheet.insert_cols(max_col + 1, amount=2)

    if unit_type =="VAV":
        # Set the titles for the added columns
        startup_sheet.cell(row=row, column=max_col + 1, value="Min CFM")
        startup_sheet.cell(row=row, column=max_col + 2, value="Max CFM")
        startup_sheet.cell(row=row, column=max_col + 3, value="Engineer Name")
        startup_sheet.cell(row=row, column=max_col + 4, value="               NOTES               ")
    else:
        # Set the titles for the added columns
        startup_sheet.cell(row=row, column=max_col + 1, value="Engineer Name")
        startup_sheet.cell(row=row, column=max_col + 2, value="               NOTES               ")

    # Print information to indicate that the operation is complete
    print("startup Initials and NOTES rows added to the very end.")


def _set_title_in_startup_sheet(startup_sheet, job_name):
    # Merge cells and set Job Name
    startup_sheet.merge_cells('A1:C1')
    cell = startup_sheet.cell(row=1, column=1)
    cell.value = "Job Name:"
    cell.alignment = Alignment(horizontal='center')

    startup_sheet.merge_cells('D1:N1')
    cell = startup_sheet.cell(row=1, column=4)
    cell.value = job_name
    cell.alignment = Alignment(horizontal='center')


def insert_image_into_sheet(image_path, sheet):
    try:
        # Load the image
        img = Image(image_path)
        img.height = img.height * .75  # Expand the image height to cover 9 rows
        start_col = get_first_empty_col(sheet, 5) - 1
        start_col_char = chr(64 + start_col)

        start_cell = start_col_char + '1'
        # Add the image to the worksheet (top right corner)
        sheet.add_image(img, start_cell)
    except Exception as e:
        print(f"Unable to add Image: {e}")


def resize_startup_sheet(sheet):
    change_all_column_width(sheet)
    # sheet.column_dimensions['D'].width = 10
    # sheet.column_dimensions['E'].width = 10
    # sheet.column_dimensions['F'].width = 10
    # sheet.column_dimensions['G'].width = 10


def add_units_to_sheet_startup(sheet, unit_type, number_of_units):
    # Assuming unit_names is a list of unit names
    for i in range(number_of_units):
        row_index = i + 6  # Start at row 2
        sheet.cell(row=row_index, column=3, value=unit_type)


def main():
    # variables that are passed in from points list reader. tthese are testing dummy vars. oh goodness i hope this works
    job_name = "test Job Name"  # title
    title = "Dac - 304 Controller Points List for EF(typ.Of 2)"
    controller_type = "Dac - 304"
    unit_type = "EF"
    number_of_units = 2
    ip_op_dict = {
        'IP1': ('Binary (Dry Contact)', 'EF-6 Status'),
        'OP1': ('Binary (24 VAC)', 'EF-6 Start-Stop'),
        'IP2': ('Univ.', 'Spare'),
        'OP2': ('Binary (24 VAC)', 'Spare'),
        'IP3': ('Univ.', 'Spare'),
        'OP3': ('Binary (24 VAC)', 'Sp1are')
    }
    target_sheet_loc = r'C:\Users\delta\PycharmProjects\Project Tracking Excel Sheet'

    # creating init workbook
    workbook, full_path = create_excel_workbook(job_name, unit_type, target_sheet_loc)

    # Setting Up workbook
    build_install_sheet(workbook, full_path, unit_type, number_of_units, ip_op_dict)

    close_workbook(workbook)


if __name__ == "__main__":
    main()