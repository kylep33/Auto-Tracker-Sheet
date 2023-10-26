import os
import sys

import openpyxl
import Points_List_Reader
import install_sheet_creator
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

def read_points_list(file_path):
    # Load the points list workbook
    points_list_workbook = Points_List_Reader.load_excel_file(file_path)

    # Get job name
    job_name = Points_List_Reader.read_job(points_list_workbook.active)

    # Read and display title
    title = Points_List_Reader.read_title(points_list_workbook.active)
    print("Title:", title)

    title_info = Points_List_Reader.parse_title_text(title)
    if title_info:
        controller_type = title_info['controller_type']
        unit_type = title_info['unit_type']
        num_of_units = title_info['num_of_units']
        print("Controller Type:", controller_type)
        print("Unit Type:", unit_type)
        print("Number of Units:", num_of_units)
    else:
        print("No match found.")
        return None

    # Extract unit type from the title
    unit_type = Points_List_Reader.extract_unit_type(title)
    print("Extracted Unit Types:", unit_type)

    # Create the IP-OP dictionary
    ip_op_dict = Points_List_Reader.create_ip_op_dict(points_list_workbook.active)

    return job_name, unit_type, num_of_units, ip_op_dict

def create_install_sheets(workbook,full_path, job_name, unit_type, num_of_units, ip_op_dict):
    install_sheet_creator.create_unit_sheets(workbook, unit_type)
    install_sheet_creator.build_workbook(workbook, full_path, job_name, unit_type, int(num_of_units), ip_op_dict)

def split_excel_sheets(input_file_path):
    file_paths = []  # To store the new file paths

    # Load the input Excel file
    input_wb = openpyxl.load_workbook(input_file_path)

    # Loop through each sheet in the input workbook
    for sheet_name in input_wb.sheetnames:
        sheet = input_wb[sheet_name]

        # Check if the sheet name contains "BOM" or "Bill of Materials"
        if "BOM" in sheet_name or "Bill of Materials" in sheet_name:
            continue  # Skip this sheet

        # Create a new workbook and copy the sheet to it
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Remove the default sheet
        new_wb.create_sheet(title=sheet_name)
        new_sheet = new_wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

        # Save the new workbook with the sheet to a new file
        output_file_path = f"{os.path.splitext(input_file_path)[0]}_{sheet_name}.xlsx"
        new_wb.save(output_file_path)

        # Append the new file path to the list
        file_paths.append(output_file_path)

    return file_paths

def find_points_list_file():
    # Get the directory of the script or the executable
    script_directory = os.path.dirname(sys.argv[0])
    points_list_file = None

    # Search for .xlsx files containing "points list" or "points lists" in the directory
    for root, _, files in os.walk(script_directory):
        for file in files:
            if file.lower().find("points list") != -1 or file.lower().find("points lists") != -1:
                points_list_file = os.path.join(root, file)
                break  # Exit the loop if a suitable file is found

    return points_list_file

def main():
    if getattr(sys, 'frozen', False):  # Check if running as an EXE
        # Get the directory of the script or executable
        target_directory = os.path.dirname(sys.executable)

        # Find the points list file in the script's directory
        target_points_list = find_points_list_file()

        if not target_points_list:
            print("Points list file not found in the script's directory.")
            return
    else:
        # Default values if not running as an EXE
        target_directory = r'C:\Users\delta\PycharmProjects\Project Tracking Excel Sheet'
        target_points_list = r"C:\Users\delta\PycharmProjects\Project Tracking Excel Sheet\testing_dir\Bill of Materials and Points List.xlsx"
    # List to store unit types
    unit_types = {}

    # List of file paths for points lists
    file_paths = split_excel_sheets(target_points_list)

    # Create a single workbook to hold all sheets
    combined_workbook = openpyxl.Workbook()
    combined_workbook.remove(combined_workbook.active)  # Remove the default sheet
    job_name = "[job name]"
    for file_path in file_paths:
        result = read_points_list(file_path)
        if result:
            job_name, unit_type, num_of_units, ip_op_dict = result

            # Check if the unit_type is already in the unit_types dictionary
            if unit_type in unit_types:
                # If it exists, increment the count and update unit_type
                unit_types[unit_type] += 1
                enumerated_unit_type = f"{unit_type} ({unit_types[unit_type]})"
                unit_type = enumerated_unit_type
            else:
                # If it's the first occurrence, add it to the dictionary
                unit_types[unit_type] = 0

            # Call the function to create the install sheets within the combined workbook
            create_install_sheets(combined_workbook, file_path, job_name, unit_type, num_of_units, ip_op_dict)

    # Save the combined workbook to a single Excel file
    combined_file_path = os.path.join(target_directory, f"Project Tracker - {job_name}.xlsx")
    combined_workbook.save(combined_file_path)

    for path in file_paths:
        os.remove(path)


if __name__ == "__main__":
    main()
